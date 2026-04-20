import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re
import io
from copy import copy

st.set_page_config(page_title="롯데마트 발주 자동화", page_icon="📦", layout="wide")

st.title("📦 롯데마트 발주 자동화")
st.markdown("서식파일과 주문파일을 업로드하면 자동으로 RAW 데이터와 수주 요약을 생성합니다.")

col1, col2 = st.columns(2)
with col1:
    template_file = st.file_uploader("📋 서식파일 업로드 (.xlsx)", type=["xlsx"], key="template")
with col2:
    order_file = st.file_uploader("📄 주문파일 업로드 (.xlsx)", type=["xlsx"], key="order")

def parse_order_file(order_bytes):
    """주문파일에서 센터별 데이터 파싱"""
    df_raw = pd.read_excel(io.BytesIO(order_bytes), header=None)
    
    centers = []
    current_center = None
    items = []
    
    # 문서번호, 발주코드 매핑 저장
    doc_info = {}
    
    i = 0
    while i < len(df_raw):
        row = df_raw.iloc[i].tolist()
        row_str = [str(v) if pd.notna(v) else '' for v in row]
        
        # 헤더행 감지 (문서명, 문서번호 등)
        if '문서명' in row_str and '점포(센터)' in row_str:
            i += 1
            data_row = df_raw.iloc[i].tolist()
            data_str = [str(v) if pd.notna(v) else '' for v in data_row]
            
            # 센터명과 발주코드 추출
            센터명 = ''
            발주코드 = ''
            문서번호 = ''
            
            try:
                center_idx = row_str.index('점포(센터)')
                센터명 = data_str[center_idx]
            except (ValueError, IndexError):
                pass
            
            try:
                doc_idx = row_str.index('문서번호')
                문서번호 = data_str[doc_idx]
            except (ValueError, IndexError):
                pass
            
            if current_center and items:
                centers.append({'센터': current_center, '발주코드': doc_info.get(current_center, ''), '문서번호': doc_info.get(current_center + '_doc', ''), 'items': items})
                items = []
            
            current_center = 센터명
            doc_info[센터명] = 발주코드
            doc_info[센터명 + '_doc'] = 문서번호
            i += 1
            continue
        
        # 상품코드 헤더행 스킵
        if '상품코드' in row_str and '판매코드' in row_str and '주문수' in row_str:
            i += 1
            continue
        
        # 합계행 스킵
        if '합계' in row_str:
            i += 1
            continue
        
        # 실제 데이터행 감지 (상품코드가 숫자 형태)
        first_val = str(row[0]) if pd.notna(row[0]) else ''
        if first_val and first_val.replace('.0', '').isdigit() and len(first_val.replace('.0', '')) >= 10:
            # 주문수에서 (BOX) 제거
            주문수_raw = str(row[6]) if pd.notna(row[6]) else '0'
            주문수 = re.sub(r'\s*\(BOX\)', '', 주문수_raw).strip()
            try:
                주문수_int = int(float(주문수))
            except:
                주문수_int = 0
            
            item = {
                '상품코드': str(row[0]).replace('.0', '') if pd.notna(row[0]) else '',
                '판매코드': str(row[1]).replace('.0', '') if pd.notna(row[1]) else '',
                '상품명': str(row[2]) if pd.notna(row[2]) else '',
                '점포명': str(row[3]) if pd.notna(row[3]) else '',
                '규격': str(row[4]) if pd.notna(row[4]) else '',
                '입수': str(row[5]).replace('.0', '') if pd.notna(row[5]) else '',
                '주문수': 주문수_int,
                '단가': str(row[7]).replace(',', '') if pd.notna(row[7]) else '',
                '주문금액': str(row[8]).replace(',', '') if pd.notna(row[8]) else '',
                '입고허용일': str(row[9]).replace('.0', '') if pd.notna(row[9]) else '',
            }
            items.append(item)
        
        i += 1
    
    # 마지막 센터 처리
    if current_center and items:
        centers.append({'센터': current_center, '발주코드': doc_info.get(current_center, ''), '문서번호': doc_info.get(current_center + '_doc', ''), 'items': items})
    
    return centers


def get_template_info(template_bytes):
    """서식파일에서 센터-배송코드 매핑 및 상품코드-발주코드 매핑 추출"""
    wb = load_workbook(io.BytesIO(template_bytes), data_only=True)
    
    # RAW 시트에서 센터-발주코드 매핑
    raw_sheet = wb['롯데마트 RAW']
    center_to_order_code = {}
    for row in raw_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:  # 점포코드, 센터명
            center_to_order_code[str(row[1])] = str(row[0])
    
    # 수주 시트에서 상품코드-발주코드-배송코드 매핑
    order_sheet = wb['롯데마트 수주']
    product_info = {}  # 상품코드 -> {발주코드, 배송코드, 센터, 품명, 단가}
    for row in order_sheet.iter_rows(min_row=2, values_only=True):
        if row[5] and row[6]:  # 상품코드, 품명
            product_info[str(row[5])] = {
                '발주코드': str(row[1]) if row[1] else '',
                '배송코드': str(row[3]) if row[3] else '',
                '센터': str(row[4]) if row[4] else '',
                '품명': str(row[6]) if row[6] else '',
                '단가': row[8] if row[8] else 0,
            }
    
    return center_to_order_code, product_info


def build_output(template_bytes, order_centers, center_to_order_code, product_info):
    """서식파일 기반으로 출력 엑셀 생성"""
    wb = load_workbook(io.BytesIO(template_bytes))
    
    # ─── 시트 1: 롯데마트 RAW ───────────────────────────────
    raw_sheet = wb['롯데마트 RAW']
    
    # 기존 데이터 삭제 (헤더 유지)
    for row in raw_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.value = None
    
    current_row = 2
    for center_data in order_centers:
        센터명 = center_data['센터']
        발주코드 = center_to_order_code.get(센터명, '')
        
        items = center_data['items']
        if not items:
            continue
        
        # 점포명을 제외하고 상품코드~입고허용일 기입
        for item in items:
            raw_sheet.cell(row=current_row, column=1).value = 발주코드  # 점포코드
            raw_sheet.cell(row=current_row, column=2).value = 센터명    # 점포(센터)
            raw_sheet.cell(row=current_row, column=3).value = item['상품코드']
            raw_sheet.cell(row=current_row, column=4).value = item['판매코드']
            raw_sheet.cell(row=current_row, column=5).value = item['상품명']
            raw_sheet.cell(row=current_row, column=6).value = item['점포명']  # 점포명
            raw_sheet.cell(row=current_row, column=7).value = item['규격']
            raw_sheet.cell(row=current_row, column=8).value = item['입수']
            raw_sheet.cell(row=current_row, column=9).value = item['주문수']  # (BOX) 제거된 값
            raw_sheet.cell(row=current_row, column=10).value = item['단가']
            raw_sheet.cell(row=current_row, column=11).value = item['주문금액']
            raw_sheet.cell(row=current_row, column=12).value = item['입고허용일']
            current_row += 1
    
    # ─── 시트 2: 롯데마트 수주 ──────────────────────────────
    order_sheet = wb['롯데마트 수주']
    
    # 기존 데이터 삭제 (헤더 유지)
    for row in order_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.value = None
    
    # 주문 데이터 집계: 센터 + 상품코드 기준으로 수량 합산
    from collections import defaultdict
    agg = defaultdict(lambda: {'수량': 0, '단가': 0, '품명': '', '발주코드': '', '배송코드': '', '센터': ''})
    
    for center_data in order_centers:
        센터명 = center_data['센터']
        발주코드 = center_to_order_code.get(센터명, '')
        
        for item in center_data['items']:
            # 상품코드로 배송코드 찾기
            prod_code = item['상품코드']
            prod_meta = None
            for code, meta in product_info.items():
                if prod_code in code or code in prod_code:
                    prod_meta = meta
                    break
            
            배송코드 = prod_meta['배송코드'] if prod_meta else ''
            품명 = prod_meta['품명'] if prod_meta else item['상품명']
            단가 = int(str(item['단가']).replace(',', '')) if item['단가'] else 0
            
            key = (발주코드, 배송코드, 센터명)
            agg[key]['수량'] += item['주문수']
            agg[key]['단가'] = 단가
            agg[key]['품명'] = 품명
            agg[key]['발주코드'] = 발주코드
            agg[key]['배송코드'] = 배송코드
            agg[key]['센터'] = 센터명
    
    # 수량 > 0인 항목만 기입
    current_row = 2
    for key, data in agg.items():
        if data['수량'] > 0:
            total = data['수량'] * data['단가']
            order_sheet.cell(row=current_row, column=2).value = data['발주코드']
            order_sheet.cell(row=current_row, column=4).value = data['배송코드']
            order_sheet.cell(row=current_row, column=5).value = data['센터']
            order_sheet.cell(row=current_row, column=6).value = key[1]  # 배송코드(상품코드)
            order_sheet.cell(row=current_row, column=7).value = data['품명']
            order_sheet.cell(row=current_row, column=8).value = data['수량']
            order_sheet.cell(row=current_row, column=9).value = data['단가']
            order_sheet.cell(row=current_row, column=10).value = total
            current_row += 1
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ─── 메인 로직 ──────────────────────────────────────────────
if template_file and order_file:
    template_bytes = template_file.read()
    order_bytes = order_file.read()
    
    with st.spinner("파일 분석 중..."):
        try:
            order_centers = parse_order_file(order_bytes)
            center_to_order_code, product_info = get_template_info(template_bytes)
        except Exception as e:
            st.error(f"파일 파싱 오류: {e}")
            st.stop()
    
    # 미리보기
    st.markdown("---")
    st.subheader("📊 파싱 결과 미리보기")
    
    for center_data in order_centers:
        with st.expander(f"🏭 {center_data['센터']} ({len(center_data['items'])}개 항목)"):
            df_preview = pd.DataFrame(center_data['items'])
            st.dataframe(df_preview, use_container_width=True)
    
    # 수주 요약 미리보기
    st.subheader("📋 수주 요약 (수량 > 0)")
    from collections import defaultdict
    agg = defaultdict(lambda: {'수량': 0, '단가': 0, '품명': '', '발주코드': '', '배송코드': '', '센터': ''})
    
    for center_data in order_centers:
        센터명 = center_data['센터']
        발주코드 = center_to_order_code.get(센터명, '')
        for item in center_data['items']:
            prod_code = item['상품코드']
            prod_meta = None
            for code, meta in product_info.items():
                if prod_code in code or code in prod_code:
                    prod_meta = meta
                    break
            배송코드 = prod_meta['배송코드'] if prod_meta else ''
            품명 = prod_meta['품명'] if prod_meta else item['상품명']
            단가 = int(str(item['단가']).replace(',', '')) if item['단가'] else 0
            key = (발주코드, 배송코드, 센터명)
            agg[key]['수량'] += item['주문수']
            agg[key]['단가'] = 단가
            agg[key]['품명'] = 품명
            agg[key]['발주코드'] = 발주코드
            agg[key]['배송코드'] = 배송코드
            agg[key]['센터'] = 센터명
    
    summary_rows = []
    for key, data in agg.items():
        if data['수량'] > 0:
            summary_rows.append({
                '발주코드': data['발주코드'],
                '배송코드': data['배송코드'],
                '센터': data['센터'],
                '상품코드': key[1],
                '품명': data['품명'],
                'Unit수량': data['수량'],
                '단가': data['단가'],
                'Total Amount': data['수량'] * data['단가'],
            })
    
    if summary_rows:
        df_summary = pd.DataFrame(summary_rows)
        st.dataframe(df_summary, use_container_width=True)
    else:
        st.info("수량 > 0인 항목이 없습니다.")
    
    # 출력 파일 생성
    st.markdown("---")
    if st.button("📥 엑셀 파일 생성 및 다운로드", type="primary"):
        with st.spinner("엑셀 생성 중..."):
            try:
                output = build_output(template_bytes, order_centers, center_to_order_code, product_info)
                st.download_button(
                    label="💾 완성된 엑셀 다운로드",
                    data=output,
                    file_name="롯데마트_발주서_완성.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                st.success("✅ 엑셀 파일 생성 완료!")
            except Exception as e:
                st.error(f"엑셀 생성 오류: {e}")

else:
    st.info("👆 위에서 두 파일을 모두 업로드해주세요.")
    st.markdown("""
    **파일 설명:**
    - **서식파일**: 롯데마트 RAW, 수주 시트가 포함된 템플릿
    - **주문파일**: 센터별 주문 데이터 (예: `라떼는.xlsx`)
    """)
